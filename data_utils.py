from docx import Document
from openpyxl import load_workbook
from io import BytesIO
import pandas as pd
from typing import List, Tuple, Dict


# ✅ docx 파일 처리 함수 (문단 추출)
def process_docx_file(uploaded_file):
    document = Document(uploaded_file)

    chunks = []
    current_chunk = []

    for para in document.paragraphs:
        text = para.text.strip()
        if not text:
            if current_chunk:
                chunks.append("\n".join(current_chunk))
                current_chunk = []
        else:
            current_chunk.append(text)

    if current_chunk:
        chunks.append("\n".join(current_chunk))

    data_list = []
    for i, chunk in enumerate(chunks):
        # 첫 라인으로 레이블 추출
        first_line = chunk.splitlines()[0] if chunk else f"paragraph_{i}"
        data_list.append(
            {
                "id": f"paragraph_{i}",
                "type": "text",
                "label": first_line,
                "content": chunk,
            }
        )
    return data_list

def find_tables(ws) -> List[Tuple[int,int,int,int]]:
    max_row = ws.max_row
    max_col = ws.max_column

    visited = [[False]*(max_col+1) for _ in range(max_row+1)]
    tables = []

    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            if visited[r][c]:
                continue
            
            is_valid_start_point = False

            current_cell_has_value = ws.cell(row=r, column=c).value is not None
            
            if r > 1 and ws.cell(row=r-1, column=c).value is not None and not visited[r-1][c]:
                 continue
            if c > 1 and ws.cell(row=r, column=c-1).value is not None and not visited[r][c-1]:
                 continue

            has_right_neighbor_value = (c + 1 <= max_col and ws.cell(row=r, column=c+1).value is not None)
            has_bottom_neighbor_value = (r + 1 <= max_row and ws.cell(row=r+1, column=c).value is not None)

            if current_cell_has_value or (has_right_neighbor_value and has_bottom_neighbor_value):
                is_valid_start_point = True
            
            if is_valid_start_point:
                start_row, start_col = r, c
                end_row, end_col = r, c

                while end_row < max_row and \
                      any(ws.cell(row=end_row+1, column=col).value is not None and not visited[end_row+1][col] \
                          for col in range(start_col, end_col+1)):
                    end_row += 1
                
                while end_col < max_col and \
                      any(ws.cell(row=row, column=end_col+1).value is not None and not visited[row][end_col+1] \
                          for row in range(start_row, end_row+1)):
                    end_col += 1

                for rr in range(start_row, end_row+1):
                    for cc in range(start_col, end_col+1):
                        visited[rr][cc] = True

                tables.append((start_row, end_row, start_col, end_col))
    
    return tables


def extract_table(ws, start_row, end_row, start_col, end_col) -> pd.DataFrame:
    """
    주어진 영역에서 테이블을 추출하여 DataFrame으로 변환.
    헤더 없이 모든 행을 데이터로 처리하고, 기본 숫자 헤더를 사용합니다.
    """
    data = []
    for r in range(start_row, end_row + 1):
        row = []
        for c in range(start_col, end_col + 1):
            row.append(ws.cell(row=r, column=c).value)
        data.append(row)

    # 헤더 없이 모든 데이터를 바로 DataFrame으로 만듭니다.
    # columns=None으로 설정하면 pandas가 기본 숫자 인덱스를 컬럼명으로 사용합니다.
    df = pd.DataFrame(data)

    return df

def should_split_first_row(row: pd.Series) -> bool:
    return pd.notna(row.iloc[0]) and row.iloc[1:].isna().all()

def split_df_on_first_row(df: pd.DataFrame) -> List[pd.DataFrame]:
    if df.empty or len(df) < 2:
        return [df]

    first_row = df.iloc[0]
    if pd.notna(first_row.iloc[0]) and first_row.iloc[1:].isna().all():
        df1 = pd.DataFrame([[first_row.iloc[0]]])
        df2 = df.iloc[1:].reset_index(drop=True)
        return [df1, df2]
    else:
        return [df]

def process_xlsx_file(uploaded_file: BytesIO) -> Dict[str, List[pd.DataFrame]]:
    wb = load_workbook(uploaded_file, data_only=True)
    sheet_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tables = find_tables(ws)
        dfs = []
        for (start_row, end_row, start_col, end_col) in tables:
            df = extract_table(ws, start_row, end_row, start_col, end_col)
            dfs.append(df)

        final_dfs = []
        for df in dfs:
            final_dfs.extend(split_df_on_first_row(df))

        if final_dfs:
            sheet_data[sheet_name] = final_dfs

    return sheet_data